from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json
import re
import uuid

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None
    get_column_letter = None


@dataclass
class XlsxBlock:
    block_id: str
    source_file: str
    source_path: str
    doc_type: str
    block_type: str
    order: int
    text: str
    section_path: List[str] = field(default_factory=list)
    page_number: Optional[int] = None
    slide_number: Optional[int] = None
    shape_index: Optional[int] = None
    sheet_name: Optional[str] = None
    row_number: Optional[int] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


class XlsxParserError(Exception):
    """Raised when an XLSX workbook cannot be parsed safely."""


class XlsxParser:
    def __init__(
        self,
        include_workbook_metadata: bool = True,
        include_sheet_metadata: bool = True,
        include_empty_rows: bool = False,
        infer_header_row: bool = True,
    ) -> None:
        self.include_workbook_metadata = include_workbook_metadata
        self.include_sheet_metadata = include_sheet_metadata
        self.include_empty_rows = include_empty_rows
        self.infer_header_row = infer_header_row

    def _make_block_id(self) -> str:
        return f"xlsx_{uuid.uuid4().hex[:12]}"

    def _normalize_text(self, text: Optional[str]) -> str:
        if text is None:
            return ""

        text = str(text).replace("\r", "\n")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()

    def _safe_scalar(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (bool, int, float)):
            return value
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        return self._normalize_text(str(value))

    def _cell_to_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        if isinstance(value, (date, time)):
            return value.isoformat()
        return self._normalize_text(str(value))

    def _column_letter(self, column_index: int) -> str:
        if get_column_letter is None:
            return f"Column {column_index}"
        return get_column_letter(column_index)

    def _dedupe_labels(self, labels: List[Tuple[int, str]]) -> Dict[int, str]:
        seen: Dict[str, int] = {}
        deduped: Dict[int, str] = {}

        for column_index, label in labels:
            base_label = self._normalize_text(label) or self._column_letter(column_index)
            seen[base_label] = seen.get(base_label, 0) + 1
            occurrence = seen[base_label]
            final_label = base_label if occurrence == 1 else f"{base_label} ({occurrence})"
            deduped[column_index] = final_label

        return deduped

    def _workbook_metadata(self, workbook: Any) -> Dict[str, Any]:
        properties = getattr(workbook, "properties", None)

        metadata = {
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": list(workbook.sheetnames),
            "creator": self._safe_scalar(getattr(properties, "creator", None)) if properties else None,
            "title": self._safe_scalar(getattr(properties, "title", None)) if properties else None,
            "subject": self._safe_scalar(getattr(properties, "subject", None)) if properties else None,
            "description": self._safe_scalar(getattr(properties, "description", None)) if properties else None,
            "keywords": self._safe_scalar(getattr(properties, "keywords", None)) if properties else None,
            "category": self._safe_scalar(getattr(properties, "category", None)) if properties else None,
            "created": self._safe_scalar(getattr(properties, "created", None)) if properties else None,
            "modified": self._safe_scalar(getattr(properties, "modified", None)) if properties else None,
            "last_modified_by": self._safe_scalar(getattr(properties, "lastModifiedBy", None)) if properties else None,
        }

        return {key: value for key, value in metadata.items() if value not in (None, "", [])}

    def _metadata_to_text(self, metadata: Dict[str, Any]) -> str:
        lines: List[str] = []

        for key, value in metadata.items():
            if value in (None, "", []):
                continue

            pretty_key = key.replace("_", " ").title()
            if isinstance(value, list):
                lines.append(f"{pretty_key}: {', '.join(str(item) for item in value)}")
            else:
                lines.append(f"{pretty_key}: {value}")

        return "\n".join(lines)

    def _extract_rows(self, worksheet: Any) -> Tuple[List[Dict[str, Any]], int]:
        parsed_rows: List[Dict[str, Any]] = []
        max_populated_columns = 0

        for fallback_row_number, row in enumerate(worksheet.iter_rows(), start=1):
            row_number = fallback_row_number
            cells = []

            for cell in row:
                value = self._cell_to_text(cell.value)
                if not value:
                    continue

                column_index = int(cell.column)
                max_populated_columns = max(max_populated_columns, column_index)

                cell_record = {
                    "ref": cell.coordinate,
                    "column_index": column_index,
                    "value": value,
                }

                if getattr(cell, "data_type", None) == "f":
                    cell_record["formula"] = value

                cells.append(cell_record)

            if cells or self.include_empty_rows:
                parsed_rows.append(
                    {
                        "row_number": row_number,
                        "cells": cells,
                    }
                )

        return parsed_rows, max_populated_columns

    def _row_to_text(
        self,
        cells: List[Dict[str, Any]],
        header_labels: Dict[int, str],
    ) -> Tuple[str, Dict[str, Any]]:
        parts: List[str] = []
        cell_map: Dict[str, Any] = {}
        formulas: Dict[str, str] = {}

        for cell in cells:
            column_index = cell["column_index"]
            label = header_labels.get(column_index, self._column_letter(column_index))
            value = cell["value"]
            parts.append(f"{label}: {value}")
            cell_map[label] = value

            if "formula" in cell:
                formulas[cell["ref"]] = cell["formula"]

        metadata: Dict[str, Any] = {
            "cell_refs": [cell["ref"] for cell in cells],
            "non_empty_cell_count": len(cells),
            "cells": cell_map,
        }
        if formulas:
            metadata["formulas"] = formulas

        return " | ".join(parts), metadata

    def parse(self, file_path: str) -> List[XlsxBlock]:
        path = Path(file_path)
        source_path = str(path.resolve())

        if load_workbook is None:
            raise XlsxParserError(
                "openpyxl is required for XLSX parsing but is not currently installed."
            )

        if not path.exists():
            raise XlsxParserError(f"XLSX file not found: {source_path}")

        if path.is_dir():
            raise XlsxParserError(f"Expected an XLSX file but received a directory: {source_path}")

        workbook = None
        try:
            workbook = load_workbook(
                filename=str(path),
                read_only=True,
                data_only=False,
            )

            blocks: List[XlsxBlock] = []
            order = 0

            if self.include_workbook_metadata:
                workbook_metadata = self._workbook_metadata(workbook)
                workbook_text = self._metadata_to_text(workbook_metadata)
                if workbook_text:
                    order += 1
                    blocks.append(
                        XlsxBlock(
                            block_id=self._make_block_id(),
                            source_file=path.name,
                            source_path=source_path,
                            doc_type="xlsx",
                            block_type="workbook_metadata",
                            order=order,
                            text=workbook_text,
                            metadata=workbook_metadata,
                        )
                    )

            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                sheet_name = worksheet.title
                rows, max_populated_columns = self._extract_rows(worksheet)

                header_labels: Dict[int, str] = {}
                header_row_number: Optional[int] = None
                populated_row_count = sum(1 for row in rows if row["cells"])
                data_row_count = populated_row_count

                if self.infer_header_row:
                    for row in rows:
                        if row["cells"]:
                            header_row_number = row["row_number"]
                            header_labels = self._dedupe_labels(
                                [
                                    (cell["column_index"], cell["value"])
                                    for cell in row["cells"]
                                ]
                            )
                            data_row_count = max(populated_row_count - 1, 0)
                            break

                if self.include_sheet_metadata:
                    order += 1
                    blocks.append(
                        XlsxBlock(
                            block_id=self._make_block_id(),
                            source_file=path.name,
                            source_path=source_path,
                            doc_type="xlsx",
                            block_type="sheet_metadata",
                            order=order,
                            text=(
                                f"Sheet: {sheet_name}\n"
                                f"Rows With Content: {populated_row_count}\n"
                                f"Data Rows: {data_row_count}\n"
                                f"Worksheet Max Row: {worksheet.max_row}\n"
                                f"Worksheet Max Column: {worksheet.max_column}\n"
                                f"Max Populated Column Index: {max_populated_columns}"
                            ),
                            section_path=[sheet_name],
                            sheet_name=sheet_name,
                            metadata={
                                "sheet_index": sheet_index,
                                "sheet_state": worksheet.sheet_state,
                                "rows_with_content": populated_row_count,
                                "data_row_count": data_row_count,
                                "worksheet_max_row": worksheet.max_row,
                                "worksheet_max_column": worksheet.max_column,
                                "max_populated_column_index": max_populated_columns,
                                "header_row_number": header_row_number,
                            },
                        )
                    )

                if header_labels:
                    ordered_headers = [
                        header_labels[column_index]
                        for column_index in sorted(header_labels)
                    ]
                    order += 1
                    blocks.append(
                        XlsxBlock(
                            block_id=self._make_block_id(),
                            source_file=path.name,
                            source_path=source_path,
                            doc_type="xlsx",
                            block_type="header",
                            order=order,
                            text="Columns: " + ", ".join(ordered_headers),
                            section_path=[sheet_name],
                            sheet_name=sheet_name,
                            row_number=header_row_number,
                            metadata={
                                "header_row_number": header_row_number,
                                "columns": ordered_headers,
                            },
                        )
                    )

                for row in rows:
                    if not row["cells"]:
                        continue
                    if header_row_number is not None and row["row_number"] == header_row_number:
                        continue

                    row_text, row_metadata = self._row_to_text(row["cells"], header_labels)
                    if not row_text:
                        continue

                    order += 1
                    blocks.append(
                        XlsxBlock(
                            block_id=self._make_block_id(),
                            source_file=path.name,
                            source_path=source_path,
                            doc_type="xlsx",
                            block_type="row",
                            order=order,
                            text=row_text,
                            section_path=[sheet_name],
                            sheet_name=sheet_name,
                            row_number=row["row_number"],
                            metadata=row_metadata,
                        )
                    )

            return blocks

        except XlsxParserError:
            raise
        except OSError as exc:
            raise XlsxParserError(f"Could not open XLSX '{source_path}': {exc}") from exc
        except Exception as exc:
            raise XlsxParserError(f"Unexpected error while parsing XLSX '{source_path}': {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()


def parse_xlsx_folder(
    folder_path: str,
    parser: Optional[XlsxParser] = None,
) -> Dict[str, List[XlsxBlock]]:
    parser = parser or XlsxParser()
    folder = Path(folder_path)
    parsed_results: Dict[str, List[XlsxBlock]] = {}

    for file_path in sorted(folder.glob("*.xlsx")):
        if file_path.name.startswith("~$"):
            continue
        parsed_results[file_path.name] = parser.parse(str(file_path))

    return parsed_results


def save_results_to_json(
    parsed_results: Dict[str, List[XlsxBlock]],
    output_file: str = "parsed_xlsx_output.json",
) -> None:
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    json_ready: Dict[str, Any] = {}

    for file_name, blocks in parsed_results.items():
        json_ready[file_name] = [asdict(block) for block in blocks]

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(json_ready, f, indent=2, ensure_ascii=False)
